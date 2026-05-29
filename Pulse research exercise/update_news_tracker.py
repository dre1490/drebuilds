"""
Pulse Research - Daily AI News Logger
Step 3: Fetch headlines AND append them to Excel automatically

HOW TO RUN:
1. Open Terminal / Command Prompt
2. Navigate to your Pulse research exercise folder
3. Run: python update_news_tracker.py
"""

import requests
from openpyxl import load_workbook
from datetime import date

# -----------------------------------------
# YOUR API KEY
# -----------------------------------------
API_KEY = "06a13e8b6988472688d82663439fc03f"

# -----------------------------------------
# SETTINGS
# -----------------------------------------
EXCEL_FILE  = "pulse_research_news_tracker.xlsx"
TOPIC       = "Artificial Intelligence"
NUM_HEADLINES = 5

# -----------------------------------------
# STEP 1 — Fetch headlines from NewsAPI
# -----------------------------------------
print(f"\n📡 Fetching top {NUM_HEADLINES} headlines for: '{TOPIC}'\n")

url = "https://newsapi.org/v2/everything"

params = {
    "q":        TOPIC,
    "language": "en",
    "sortBy":   "publishedAt",
    "pageSize": NUM_HEADLINES,
    "apiKey":   API_KEY,
}

response = requests.get(url, params=params)
data     = response.json()

if data.get("status") != "ok":
    print(f"❌ API Error: {data.get('message', 'Unknown error')}")
    print("Check that your API key is correct.\n")
    exit()

articles = data["articles"]
today    = date.today().strftime("%Y-%m-%d")

for article in articles:
    print(f"  ✅ {article['source']['name']:<20} {article['title'][:50]}...")

# -----------------------------------------
# STEP 2 — Open Excel and append new rows
# Key difference from fund tracker:
# We ADD new rows instead of overwriting cells
# This builds a running history over time
# -----------------------------------------
print(f"\n📂 Opening {EXCEL_FILE}...")

try:
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Find the next empty row
    next_row = ws.max_row + 1

    # Remove placeholder row if it's still there
    if ws.cell(row=5, column=2).value == "Headline will appear here...":
        ws.delete_rows(5)
        next_row = 5

    # Append each headline as a new row
    for article in articles:
        headline = article["title"]
        source   = article["source"]["name"]
        link     = article["url"]

        ws.cell(row=next_row, column=1, value=today)
        ws.cell(row=next_row, column=2, value=headline)
        ws.cell(row=next_row, column=3, value=source)
        ws.cell(row=next_row, column=4, value=link)

        # Wrap text for long headlines
        ws.cell(row=next_row, column=2).alignment = \
            __import__('openpyxl').styles.Alignment(wrap_text=True)
        ws.row_dimensions[next_row].height = 30

        next_row += 1

    wb.save(EXCEL_FILE)
    print(f"\n✅ Excel file updated successfully!")
    print(f"📅 Date: {today}")
    print(f"📰 Headlines added: {len(articles)}")
    print(f"💾 Saved: {EXCEL_FILE}\n")

except FileNotFoundError:
    print(f"\n❌ Could not find {EXCEL_FILE}")
    print("Make sure it's in the same folder as this script.\n")

except Exception as e:
    print(f"\n❌ Something went wrong: {str(e)}\n")
