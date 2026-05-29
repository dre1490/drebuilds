"""
Vertex Solutions - Report Generator with Email Alerts
Upgrade 1: Get notified if there are unpaid accounts or revenue issues

HOW TO RUN:
1. Open Terminal / Command Prompt
2. Navigate to your SQL project folder
3. Run: python generate_report_with_alerts.py
"""

import sqlite3
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

# -----------------------------------------
# MAILTRAP SETTINGS
# -----------------------------------------
SMTP_HOST     = "sandbox.smtp.mailtrap.io"
SMTP_PORT     = 2525
SMTP_USERNAME = "6cba633f05a265"
SMTP_PASSWORD = "55ffb250f367b5"
EMAIL_FROM    = "reports@vertexsolutions.com"
EMAIL_TO      = "manager@vertexsolutions.com"

# -----------------------------------------
# SETTINGS
# -----------------------------------------
DB_FILE           = "vertex_solutions.db"
REPORT_FILE       = f"vertex_report_{date.today().strftime('%Y-%m-%d')}.xlsx"
MIN_REVENUE       = 1500.00  # Alert if total revenue drops below this

# Styling helpers
dark_blue  = "1F3864"
mid_blue   = "2E75B6"
light_blue = "DCE6F1"
red        = "C00000"
yellow     = "FFF2CC"
white      = "FFFFFF"

thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(cell, bold=False, size=11, color="000000", fill=None, align="left", wrap=False):
    cell.font      = Font(name="Arial", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border    = border
    if fill:
        cell.fill  = PatternFill("solid", start_color=fill)

# -----------------------------------------
# STEP 1 — Pull data from database
# -----------------------------------------
print("\n🗄️  Reading database...")

alerts = []

try:
    conn   = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    # Total customers
    cursor.execute("SELECT COUNT(*) FROM customers WHERE signup_date LIKE '2026-05%'")
    total_customers = cursor.fetchone()[0]

    # Customers by plan
    cursor.execute("""
        SELECT plan, COUNT(*) as total
        FROM customers
        GROUP BY plan
        ORDER BY total DESC
    """)
    by_plan = cursor.fetchall()

    # Revenue by plan
    cursor.execute("""
        SELECT plan, COUNT(*) as customers, SUM(monthly_fee) as revenue
        FROM customers
        GROUP BY plan
        ORDER BY revenue DESC
    """)
    revenue     = cursor.fetchall()
    total_revenue = sum(row[2] for row in revenue)

    # Unpaid customers
    cursor.execute("""
        SELECT name, email, plan, monthly_fee
        FROM customers
        WHERE paid = 'No'
        ORDER BY monthly_fee DESC
    """)
    unpaid       = cursor.fetchall()
    unpaid_total = sum(row[3] for row in unpaid)

    conn.close()
    print("✅ Data pulled successfully")

    # -----------------------------------------
    # CHECK FOR ALERT CONDITIONS
    # -----------------------------------------
    if unpaid:
        alerts.append(f"⚠️  {len(unpaid)} unpaid account(s) totaling ${unpaid_total:,.2f}")
        for name, email, plan, fee in unpaid:
            alerts.append(f"   • {name} ({plan}) — ${fee:,.2f} outstanding")

    if total_revenue < MIN_REVENUE:
        alerts.append(f"📉 Total revenue ${total_revenue:,.2f} is below threshold of ${MIN_REVENUE:,.2f}")

except FileNotFoundError:
    alerts.append("❌ Database file not found — vertex_solutions.db is missing")
    print("❌ Database not found")
    exit()

except Exception as e:
    alerts.append(f"❌ Database error: {str(e)}")
    print(f"❌ Error: {str(e)}")
    exit()

# -----------------------------------------
# STEP 2 — Build Excel report
# -----------------------------------------
print("📊 Building Excel report...")

wb = Workbook()
ws = wb.active
ws.title = "Weekly Report"

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18

today = date.today().strftime("%B %d, %Y")

# Header
ws.merge_cells("A1:D1")
ws["A1"] = "VERTEX SOLUTIONS — Weekly Customer Report"
style(ws["A1"], bold=True, size=14, color=white, fill=dark_blue, align="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:D2")
ws["A2"] = f"Generated: {today}"
style(ws["A2"], size=10, color=white, fill=mid_blue, align="center")
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 10

# Summary
ws.merge_cells("A4:D4")
ws["A4"] = "SUMMARY"
style(ws["A4"], bold=True, size=11, color=white, fill=mid_blue)
ws.row_dimensions[4].height = 20

summary_rows = [
    ("Total Customers This Month", total_customers),
    ("Total Monthly Revenue",      f"${total_revenue:,.2f}"),
    ("Outstanding Payments",       f"${unpaid_total:,.2f}"),
    ("Unpaid Accounts",            len(unpaid)),
]

for i, (label, value) in enumerate(summary_rows):
    row  = i + 5
    fill = light_blue if i % 2 == 0 else white
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=value)
    style(ws.cell(row=row, column=1), bold=True, fill=fill)
    style(ws.cell(row=row, column=2), fill=fill, align="right")
    for col in [3, 4]:
        style(ws.cell(row=row, column=col), fill=fill)
    ws.row_dimensions[row].height = 18

ws.row_dimensions[9].height = 10

# Revenue by plan
ws.merge_cells("A10:D10")
ws["A10"] = "REVENUE BY PLAN"
style(ws["A10"], bold=True, size=11, color=white, fill=mid_blue)
ws.row_dimensions[10].height = 20

headers = ["Plan", "Customers", "Monthly Revenue", "% of Total"]
for col, h in enumerate(headers, 1):
    ws.cell(row=11, column=col, value=h)
    style(ws.cell(row=11, column=col), bold=True, color=white, fill=dark_blue, align="center")
ws.row_dimensions[11].height = 18

for i, (plan, customers, rev) in enumerate(revenue):
    row  = i + 12
    pct  = (rev / total_revenue * 100) if total_revenue else 0
    fill = light_blue if i % 2 == 0 else white
    ws.cell(row=row, column=1, value=plan)
    ws.cell(row=row, column=2, value=customers)
    ws.cell(row=row, column=3, value=f"${rev:,.2f}")
    ws.cell(row=row, column=4, value=f"{pct:.1f}%")
    for col in range(1, 5):
        style(ws.cell(row=row, column=col), fill=fill, align="center")
    ws.row_dimensions[row].height = 18

ws.row_dimensions[15].height = 10

# Unpaid accounts
ws.merge_cells("A16:D16")
ws["A16"] = "⚠️  UNPAID ACCOUNTS — ACTION REQUIRED"
style(ws["A16"], bold=True, size=11, color=white, fill=red)
ws.row_dimensions[16].height = 20

headers = ["Customer Name", "Email", "Plan", "Amount Due"]
for col, h in enumerate(headers, 1):
    ws.cell(row=17, column=col, value=h)
    style(ws.cell(row=17, column=col), bold=True, color=white, fill=dark_blue, align="center")
ws.row_dimensions[17].height = 18

for i, (name, email, plan, fee) in enumerate(unpaid):
    row = i + 18
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=email)
    ws.cell(row=row, column=3, value=plan)
    ws.cell(row=row, column=4, value=f"${fee:,.2f}")
    for col in range(1, 5):
        style(ws.cell(row=row, column=col), fill=yellow, align="center")
    ws.row_dimensions[row].height = 18

total_row = 18 + len(unpaid)
ws.cell(row=total_row, column=3, value="TOTAL OUTSTANDING")
ws.cell(row=total_row, column=4, value=f"${unpaid_total:,.2f}")
style(ws.cell(row=total_row, column=3), bold=True, fill=red, color=white, align="center")
style(ws.cell(row=total_row, column=4), bold=True, fill=red, color=white, align="center")
for col in [1, 2]:
    style(ws.cell(row=total_row, column=col), fill=red)
ws.row_dimensions[total_row].height = 20

wb.save(REPORT_FILE)
print(f"✅ Report saved as: {REPORT_FILE}")

# -----------------------------------------
# STEP 3 — Send alert email if needed
# -----------------------------------------
if alerts:
    print(f"\n📧 Issues detected — sending alert email...")

    today_str = date.today().strftime("%B %d, %Y")
    subject   = f"⚠️ Vertex Solutions Weekly Report — Action Required {today_str}"

    body = f"""
Hello,

The weekly customer report ran on {today_str} and requires your attention:

{chr(10).join(alerts)}

The full Excel report has been saved as:
{REPORT_FILE}

Please review and take action on the flagged items.

— Vertex Solutions Automated Reporting System
    """

    try:
        msg            = MIMEMultipart()
        msg["From"]    = EMAIL_FROM
        msg["To"]      = EMAIL_TO
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, EMAIL_TO, msg.as_string())

        print(f"✅ Alert email sent!")
        print(f"📬 Check your Mailtrap inbox to see it!\n")

    except Exception as e:
        print(f"❌ Email failed: {str(e)}\n")

else:
    print(f"\n✅ No issues detected — no alert needed!")
    print(f"📅 Date: {date.today().strftime('%Y-%m-%d')}\n")
