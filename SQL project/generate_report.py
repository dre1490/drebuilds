"""
Vertex Solutions - Automated Weekly Report Generator
Step 3: Pull data from database and generate a clean Excel report

HOW TO RUN:
1. Open Terminal / Command Prompt
2. Navigate to your SQL project folder
3. Run: python generate_report.py
"""

import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

# -----------------------------------------
# SETTINGS
# -----------------------------------------
DB_FILE     = "vertex_solutions.db"
REPORT_FILE = f"vertex_report_{date.today().strftime('%Y-%m-%d')}.xlsx"

# Colors
dark_blue   = "1F3864"
mid_blue    = "2E75B6"
light_blue  = "DCE6F1"
red         = "C00000"
green       = "70AD47"
white       = "FFFFFF"
yellow      = "FFF2CC"

thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(cell, bold=False, size=11, color="000000", fill=None, align="left", wrap=False):
    cell.font      = Font(name="Arial", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border    = border
    if fill:
        cell.fill = PatternFill("solid", start_color=fill)

# -----------------------------------------
# STEP 1 — Pull data from database
# -----------------------------------------
print("\n🗄️  Reading database...")

conn   = sqlite3.connect(DB_FILE)
cursor = conn.cursor()

# Query 1 — Total customers
cursor.execute("SELECT COUNT(*) FROM customers WHERE signup_date LIKE '2026-05%'")
total_customers = cursor.fetchone()[0]

# Query 2 — Customers by plan
cursor.execute("""
    SELECT plan, COUNT(*) as total
    FROM customers
    GROUP BY plan
    ORDER BY total DESC
""")
by_plan = cursor.fetchall()

# Query 3 — Revenue by plan
cursor.execute("""
    SELECT plan, COUNT(*) as customers, SUM(monthly_fee) as revenue
    FROM customers
    GROUP BY plan
    ORDER BY revenue DESC
""")
revenue = cursor.fetchall()
total_revenue = sum(row[2] for row in revenue)

# Query 4 — Unpaid customers
cursor.execute("""
    SELECT name, email, plan, monthly_fee
    FROM customers
    WHERE paid = 'No'
    ORDER BY monthly_fee DESC
""")
unpaid = cursor.fetchall()
unpaid_total = sum(row[3] for row in unpaid)

conn.close()
print("✅ Data pulled successfully")

# -----------------------------------------
# STEP 2 — Build the Excel report
# -----------------------------------------
print("📊 Building Excel report...")

wb = Workbook()
ws = wb.active
ws.title = "Weekly Report"

# Column widths
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18

today = date.today().strftime("%B %d, %Y")

# -----------------------------------------
# HEADER
# -----------------------------------------
ws.merge_cells("A1:D1")
ws["A1"] = "VERTEX SOLUTIONS — Weekly Customer Report"
style(ws["A1"], bold=True, size=14, color=white, fill=dark_blue, align="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:D2")
ws["A2"] = f"Generated: {today}"
style(ws["A2"], size=10, color=white, fill=mid_blue, align="center")
ws.row_dimensions[2].height = 20

ws.row_dimensions[3].height = 10

# -----------------------------------------
# SECTION 1 — Summary
# -----------------------------------------
ws.merge_cells("A4:D4")
ws["A4"] = "SUMMARY"
style(ws["A4"], bold=True, size=11, color=white, fill=mid_blue, align="left")
ws.row_dimensions[4].height = 20

summary_rows = [
    ("Total Customers This Month", total_customers),
    ("Total Monthly Revenue",      f"${total_revenue:,.2f}"),
    ("Outstanding Payments",       f"${unpaid_total:,.2f}"),
    ("Unpaid Accounts",            len(unpaid)),
]

for i, (label, value) in enumerate(summary_rows):
    row  = i + 5
    fill = light_blue if i % 2 == 0 else white
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=value)
    style(ws.cell(row=row, column=1), bold=True, fill=fill)
    style(ws.cell(row=row, column=2), fill=fill, align="right")
    for col in [3, 4]:
        style(ws.cell(row=row, column=col), fill=fill)
    ws.row_dimensions[row].height = 18

ws.row_dimensions[9].height = 10

# -----------------------------------------
# SECTION 2 — Revenue by Plan
# -----------------------------------------
ws.merge_cells("A10:D10")
ws["A10"] = "REVENUE BY PLAN"
style(ws["A10"], bold=True, size=11, color=white, fill=mid_blue)
ws.row_dimensions[10].height = 20

headers = ["Plan", "Customers", "Monthly Revenue", "% of Total"]
for col, h in enumerate(headers, 1):
    ws.cell(row=11, column=col, value=h)
    style(ws.cell(row=11, column=col), bold=True, color=white, fill=dark_blue, align="center")
ws.row_dimensions[11].height = 18

for i, (plan, customers, rev) in enumerate(revenue):
    row  = i + 12
    pct  = (rev / total_revenue * 100) if total_revenue else 0
    fill = light_blue if i % 2 == 0 else white
    ws.cell(row=row, column=1, value=plan)
    ws.cell(row=row, column=2, value=customers)
    ws.cell(row=row, column=3, value=f"${rev:,.2f}")
    ws.cell(row=row, column=4, value=f"{pct:.1f}%")
    for col in range(1, 5):
        style(ws.cell(row=row, column=col), fill=fill, align="center")
    ws.row_dimensions[row].height = 18

ws.row_dimensions[15].height = 10

# -----------------------------------------
# SECTION 3 — Unpaid Accounts
# -----------------------------------------
ws.merge_cells("A16:D16")
ws["A16"] = "⚠️  UNPAID ACCOUNTS — ACTION REQUIRED"
style(ws["A16"], bold=True, size=11, color=white, fill=red)
ws.row_dimensions[16].height = 20

headers = ["Customer Name", "Email", "Plan", "Amount Due"]
for col, h in enumerate(headers, 1):
    ws.cell(row=17, column=col, value=h)
    style(ws.cell(row=17, column=col), bold=True, color=white, fill=dark_blue, align="center")
ws.row_dimensions[17].height = 18

for i, (name, email, plan, fee) in enumerate(unpaid):
    row = i + 18
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=email)
    ws.cell(row=row, column=3, value=plan)
    ws.cell(row=row, column=4, value=f"${fee:,.2f}")
    for col in range(1, 5):
        style(ws.cell(row=row, column=col), fill=yellow, align="center")
    ws.row_dimensions[row].height = 18

# Total row
total_row = 18 + len(unpaid)
ws.cell(row=total_row, column=3, value="TOTAL OUTSTANDING")
ws.cell(row=total_row, column=4, value=f"${unpaid_total:,.2f}")
style(ws.cell(row=total_row, column=3), bold=True, fill=red, color=white, align="center")
style(ws.cell(row=total_row, column=4), bold=True, fill=red, color=white, align="center")
for col in [1, 2]:
    style(ws.cell(row=total_row, column=col), fill=red)
ws.row_dimensions[total_row].height = 20

# -----------------------------------------
# STEP 3 — Save the report
# -----------------------------------------
wb.save(REPORT_FILE)

print(f"✅ Report generated successfully!")
print(f"📁 Saved as: {REPORT_FILE}\n")
