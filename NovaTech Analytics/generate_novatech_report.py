"""
NovaTech Analytics Platform
Phase 3: Generate a professional Excel report from the database

HOW TO RUN:
1. Open Terminal / Command Prompt
2. Navigate to your NovaTech Analytics folder
3. Run: python generate_novatech_report.py
"""

import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

# -----------------------------------------
# SETTINGS
# -----------------------------------------
DB_FILE     = "novatech.db"
REPORT_FILE = f"novatech_report_{date.today().strftime('%Y-%m-%d')}.xlsx"

# Colors
dark_blue  = "1F3864"
mid_blue   = "2E75B6"
light_blue = "DCE6F1"
green      = "70AD47"
red        = "C00000"
orange     = "ED7D31"
yellow     = "FFF2CC"
white      = "FFFFFF"

thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(cell, bold=False, size=11, color="000000", fill=None, align="left", wrap=False):
    cell.font      = Font(name="Arial", bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border    = border
    if fill:
        cell.fill  = PatternFill("solid", start_color=fill)

def section_header(ws, row, title, color=None):
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = title
    style(ws[f"A{row}"], bold=True, size=11, color=white, fill=color or mid_blue)
    ws.row_dimensions[row].height = 22

# -----------------------------------------
# PULL DATA
# -----------------------------------------
print("\n🗄️  Reading database...")

conn   = sqlite3.connect(DB_FILE)
cursor = conn.cursor()

cursor.execute("SELECT SUM(monthly_fee) FROM customers WHERE status = 'Active'")
mrr = cursor.fetchone()[0] or 0

cursor.execute("SELECT COUNT(*) FROM customers")
total_customers = cursor.fetchone()[0]

cursor.execute("SELECT COUNT(*) FROM customers WHERE status = 'Churned'")
churned = cursor.fetchone()[0]
churn_rate = (churned / total_customers * 100) if total_customers else 0

cursor.execute("""
    SELECT plan, COUNT(*) as customers, SUM(monthly_fee) as revenue
    FROM customers WHERE status = 'Active'
    GROUP BY plan ORDER BY revenue DESC
""")
by_plan = cursor.fetchall()

cursor.execute("""
    SELECT region, COUNT(*) as customers, SUM(monthly_fee) as revenue
    FROM customers WHERE status = 'Active'
    GROUP BY region ORDER BY revenue DESC
""")
by_region = cursor.fetchall()

cursor.execute("""
    SELECT product, COUNT(*) as units, SUM(amount) as revenue
    FROM sales GROUP BY product ORDER BY revenue DESC LIMIT 5
""")
top_products = cursor.fetchall()

cursor.execute("SELECT COUNT(*) FROM support_tickets")
total_tickets = cursor.fetchone()[0]

cursor.execute("SELECT COUNT(*) FROM support_tickets WHERE status = 'Open'")
open_tickets = cursor.fetchone()[0]

cursor.execute("SELECT AVG(resolution_hours) FROM support_tickets WHERE status = 'Resolved'")
avg_resolution = cursor.fetchone()[0] or 0

cursor.execute("SELECT AVG(satisfaction) FROM support_tickets WHERE satisfaction IS NOT NULL")
avg_satisfaction = cursor.fetchone()[0] or 0

cursor.execute("SELECT SUM(amount) FROM sales")
total_sales = cursor.fetchone()[0] or 0

cursor.execute("SELECT COUNT(*) FROM sales")
total_transactions = cursor.fetchone()[0]

conn.close()
print("✅ Data pulled successfully")

# -----------------------------------------
# BUILD EXCEL REPORT
# -----------------------------------------
print("📊 Building Excel report...")

wb = Workbook()
ws = wb.active
ws.title = "NovaTech Report"

# Column widths
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 18

today = date.today().strftime("%B %d, %Y")

# -----------------------------------------
# TITLE
# -----------------------------------------
ws.merge_cells("A1:F1")
ws["A1"] = "NOVATECH ANALYTICS — Executive Business Report"
style(ws["A1"], bold=True, size=16, color=white, fill=dark_blue, align="center")
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:F2")
ws["A2"] = f"Generated: {today}  |  Data Source: NovaTech Database"
style(ws["A2"], size=10, color=white, fill=mid_blue, align="center")
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 10

# -----------------------------------------
# SECTION 1 — KEY METRICS SUMMARY
# -----------------------------------------
section_header(ws, 4, "📊  KEY METRICS SUMMARY")
ws.row_dimensions[4].height = 22

metrics = [
    ("Monthly Recurring Revenue (MRR)", f"${mrr:,.2f}",         green),
    ("Total Customers",                  total_customers,         mid_blue),
    ("Churned Customers",                churned,                 red),
    ("Churn Rate",                       f"{churn_rate:.1f}%",   red if churn_rate > 7 else green),
    ("Total Sales Revenue",              f"${total_sales:,.2f}", green),
    ("Total Transactions",               total_transactions,      mid_blue),
]

for i, (label, value, color) in enumerate(metrics):
    row  = i + 5
    fill = light_blue if i % 2 == 0 else white
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=value)
    style(ws.cell(row=row, column=1), bold=True, fill=fill)
    style(ws.cell(row=row, column=2), bold=True, color=color, fill=fill, align="right")
    for col in range(3, 7):
        style(ws.cell(row=row, column=col), fill=fill)
    ws.row_dimensions[row].height = 18

ws.row_dimensions[11].height = 10

# -----------------------------------------
# SECTION 2 — REVENUE BY PLAN
# -----------------------------------------
section_header(ws, 12, "💰  REVENUE BY PLAN")

headers = ["Plan", "Active Customers", "Monthly Revenue", "% of MRR"]
for col, h in enumerate(headers, 1):
    ws.cell(row=13, column=col, value=h)
    style(ws.cell(row=13, column=col), bold=True, color=white, fill=dark_blue, align="center")
ws.row_dimensions[13].height = 18

for i, (plan, customers, revenue) in enumerate(by_plan):
    row  = i + 14
    pct  = (revenue / mrr * 100) if mrr else 0
    fill = light_blue if i % 2 == 0 else white
    ws.cell(row=row, column=1, value=plan)
    ws.cell(row=row, column=2, value=customers)
    ws.cell(row=row, column=3, value=f"${revenue:,.2f}")
    ws.cell(row=row, column=4, value=f"{pct:.1f}%")
    for col in range(1, 7):
        style(ws.cell(row=row, column=col), fill=fill, align="center")
    ws.row_dimensions[row].height = 18

ws.row_dimensions[17].height = 10

# -----------------------------------------
# SECTION 3 — REVENUE BY REGION
# -----------------------------------------
section_header(ws, 18, "🌍  REVENUE BY REGION")

headers = ["Region", "Customers", "Monthly Revenue", "% of MRR"]
for col, h in enumerate(headers, 1):
    ws.cell(row=19, column=col, value=h)
    style(ws.cell(row=19, column=col), bold=True, color=white, fill=dark_blue, align="center")
ws.row_dimensions[19].height = 18

for i, (region, customers, revenue) in enumerate(by_region):
    row  = i + 20
    pct  = (revenue / mrr * 100) if mrr else 0
    fill = light_blue if i % 2 == 0 else white
    ws.cell(row=row, column=1, value=region)
    ws.cell(row=row, column=2, value=customers)
    ws.cell(row=row, column=3, value=f"${revenue:,.2f}")
    ws.cell(row=row, column=4, value=f"{pct:.1f}%")
    for col in range(1, 7):
        style(ws.cell(row=row, column=col), fill=fill, align="center")
    ws.row_dimensions[row].height = 18

ws.row_dimensions[24].height = 10

# -----------------------------------------
# SECTION 4 — TOP PRODUCTS
# -----------------------------------------
section_header(ws, 25, "🏆  TOP SELLING PRODUCTS")

headers = ["Product", "Units Sold", "Total Revenue", "Avg Per Sale"]
for col, h in enumerate(headers, 1):
    ws.cell(row=26, column=col, value=h)
    style(ws.cell(row=26, column=col), bold=True, color=white, fill=dark_blue, align="center")
ws.row_dimensions[26].height = 18

for i, (product, units, revenue) in enumerate(top_products):
    row     = i + 27
    avg     = revenue / units if units else 0
    fill    = light_blue if i % 2 == 0 else white
    ws.cell(row=row, column=1, value=product)
    ws.cell(row=row, column=2, value=units)
    ws.cell(row=row, column=3, value=f"${revenue:,.2f}")
    ws.cell(row=row, column=4, value=f"${avg:,.2f}")
    for col in range(1, 7):
        style(ws.cell(row=row, column=col), fill=fill, align="center")
    ws.row_dimensions[row].height = 18

ws.row_dimensions[32].height = 10

# -----------------------------------------
# SECTION 5 — SUPPORT SUMMARY
# -----------------------------------------
churn_flag    = "⚠️ HIGH" if churn_rate > 7 else "✅ GOOD"
support_flag  = "⚠️ SLOW" if avg_resolution > 24 else "✅ GOOD"
sat_flag      = "⚠️ LOW"  if avg_satisfaction < 4 else "✅ GOOD"

section_header(ws, 33, "🎫  SUPPORT & SATISFACTION", color=orange)

support_metrics = [
    ("Total Support Tickets",        total_tickets,                   mid_blue),
    ("Open Tickets",                 open_tickets,                    red if open_tickets > 10 else green),
    ("Avg Resolution Time",          f"{avg_resolution:.1f} hours",   red if avg_resolution > 24 else green),
    ("Avg Satisfaction Score",       f"{avg_satisfaction:.1f} / 5.0", red if avg_satisfaction < 4 else green),
]

for i, (label, value, color) in enumerate(support_metrics):
    row  = i + 34
    fill = light_blue if i % 2 == 0 else white
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=value)
    style(ws.cell(row=row, column=1), bold=True, fill=fill)
    style(ws.cell(row=row, column=2), bold=True, color=color, fill=fill, align="right")
    for col in range(3, 7):
        style(ws.cell(row=row, column=col), fill=fill)
    ws.row_dimensions[row].height = 18

# -----------------------------------------
# SAVE
# -----------------------------------------
wb.save(REPORT_FILE)
print(f"✅ Report generated successfully!")
print(f"📁 Saved as: {REPORT_FILE}\n")
