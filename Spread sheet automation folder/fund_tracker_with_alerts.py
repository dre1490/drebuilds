"""
Horizon Capital - Fund Tracker with Email Alerts
Upgrade 1: Get notified if any price looks wrong

HOW TO RUN:
1. Open Terminal / Command Prompt
2. Navigate to your Dre AI practice file folder
3. Run: python fund_tracker_with_alerts.py
"""

import yfinance as yf
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
from datetime import date

# -----------------------------------------
# MAILTRAP SETTINGS
# Your fake email inbox credentials
# -----------------------------------------
SMTP_HOST     = "sandbox.smtp.mailtrap.io"
SMTP_PORT     = 2525
SMTP_USERNAME = "6cba633f05a265"
SMTP_PASSWORD = "55ffb250f367b5"
EMAIL_FROM    = "tracker@horizoncapital.com"
EMAIL_TO      = "dre@horizoncapital.com"

# -----------------------------------------
# SETTINGS
# -----------------------------------------
EXCEL_FILE    = "horizon_capital_tracker.xlsx"

# Alert if price is below this threshold
# Protects against bad data returning $0
MIN_PRICE     = 1.00

# -----------------------------------------
# FUND LIST
# -----------------------------------------
funds = [
    ("Vanguard 500 Index Fund",         "VFIAX", 5),
    ("Fidelity Contrafund",             "FCNTX", 6),
    ("T. Rowe Price Blue Chip Growth",  "TRBCX", 7),
    ("American Funds Growth Fund",      "AGTHX", 8),
    ("Schwab Total Stock Market Index", "SWTSX", 9),
]

# -----------------------------------------
# STEP 1 — Fetch prices
# -----------------------------------------
print("\n📡 Fetching live prices...\n")

today    = date.today().strftime("%Y-%m-%d")
results  = []
alerts   = []

for name, ticker, row in funds:
    try:
        data  = yf.Ticker(ticker)
        price = data.fast_info.get("lastPrice") or data.fast_info.get("previousClose")

        if not price:
            print(f"  ⚠️  {ticker:<8} Price not found")
            alerts.append(f"⚠️ {name} ({ticker}) — Price could not be fetched")
        elif price < MIN_PRICE:
            print(f"  ⚠️  {ticker:<8} Suspicious price: ${price:.2f}")
            alerts.append(f"⚠️ {name} ({ticker}) — Suspicious price returned: ${price:.2f}")
        else:
            print(f"  ✅ {ticker:<8} ${price:,.2f}")
            results.append((name, ticker, price, today, row))

    except Exception as e:
        print(f"  ❌ {ticker:<8} Error: {str(e)}")
        alerts.append(f"❌ {name} ({ticker}) — Script error: {str(e)}")

# -----------------------------------------
# STEP 2 — Update Excel
# -----------------------------------------
print(f"\n📂 Opening {EXCEL_FILE}...")

try:
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for name, ticker, price, date_str, row in results:
        ws.cell(row=row, column=3, value=price)
        ws.cell(row=row, column=4, value=date_str)
        ws.cell(row=row, column=5, value="✅ Updated")

    wb.save(EXCEL_FILE)
    print(f"✅ Excel updated successfully!")

except FileNotFoundError:
    alerts.append("❌ Excel file not found — make sure it's in the same folder")
    print(f"❌ Could not find {EXCEL_FILE}")

except Exception as e:
    alerts.append(f"❌ Excel error: {str(e)}")
    print(f"❌ Excel error: {str(e)}")

# -----------------------------------------
# STEP 3 — Send email alert if needed
# Only sends if something went wrong
# -----------------------------------------
if alerts:
    print(f"\n📧 Issues detected — sending alert email...")

    subject = f"⚠️ Horizon Capital Fund Tracker — Alert {today}"

    body = f"""
Hello,

The daily fund tracker ran on {today} and detected the following issues:

{chr(10).join(alerts)}

Please review the tracker and verify prices manually for the flagged funds.

— Horizon Capital Automated Tracker
    """

    try:
        msg = MIMEMultipart()
        msg["From"]    = EMAIL_FROM
        msg["To"]      = EMAIL_TO
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, EMAIL_TO, msg.as_string())

        print(f"✅ Alert email sent to {EMAIL_TO}")
        print(f"📬 Check your Mailtrap inbox to see it!\n")

    except Exception as e:
        print(f"❌ Email failed: {str(e)}\n")

else:
    print(f"\n✅ All prices look good — no alerts needed!")
    print(f"📅 Date: {today}\n")
