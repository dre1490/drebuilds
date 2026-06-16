"""
Horizon Capital - Automated Fund Tracker
Step 3: Fetch prices AND update Excel automatically

HOW TO RUN:
1. Open Terminal / Command Prompt
2. Navigate to your folder
3. Run: python update_tracker.py
"""

import yfinance as yf
from openpyxl import load_workbook
from datetime import date

# -----------------------------------------
# SETTINGS — edit these if filenames change
# -----------------------------------------
EXCEL_FILE = "horizon_capital_tracker.xlsx"
FIRST_DATA_ROW = 5  # Row where fund data starts in Excel

# -----------------------------------------
# FUND LIST — (Fund Name, Ticker, Excel Row)
# -----------------------------------------
funds = [
    ("Vanguard 500 Index Fund",         "VFIAX", 5),
    ("Fidelity Contrafund",             "FCNTX", 6),
    ("T. Rowe Price Blue Chip Growth",  "TRBCX", 7),
    ("American Funds Growth Fund",      "AGTHX", 8),
    ("Schwab Total Stock Market Index", "SWTSX", 9),
]

# -----------------------------------------
# Column layout (updated spreadsheet)
# C = Current Price (NAV)
# D = Previous Price (NAV)
# E = Date Updated
# F = Status
# -----------------------------------------
COL_CURRENT  = 3
COL_PREVIOUS = 4
COL_DATE     = 5
COL_STATUS   = 6

# -----------------------------------------
# STEP 1 — Fetch prices from Yahoo Finance
# -----------------------------------------
print("\n📡 Fetching live prices...\n")

today = date.today().strftime("%Y-%m-%d")
results = []

for name, ticker, row in funds:
    try:
        data  = yf.Ticker(ticker)
        price = data.fast_info.get("lastPrice") or data.fast_info.get("previousClose")

        if price:
            results.append((name, ticker, price, today, row))
            print(f"  ✅ {ticker:<8} ${price:,.2f}")
        else:
            print(f"  ⚠️  {ticker:<8} Price not found — skipping")

    except Exception as e:
        print(f"  ❌ {ticker:<8} Error: {str(e)}")

# -----------------------------------------
# STEP 2 — Open Excel and update the cells
# -----------------------------------------
print(f"\n📂 Opening {EXCEL_FILE}...")

try:
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for name, ticker, price, date_str, row in results:
        # Shift today's current price → previous before overwriting
        current_price = ws.cell(row=row, column=COL_CURRENT).value
        ws.cell(row=row, column=COL_PREVIOUS, value=current_price)  # Previous Price

        ws.cell(row=row, column=COL_CURRENT,  value=price)          # Current Price
        ws.cell(row=row, column=COL_DATE,     value=date_str)        # Date Updated
        ws.cell(row=row, column=COL_STATUS,   value="✅ Updated")    # Status

    wb.save(EXCEL_FILE)
    print(f"\n✅ Excel file updated successfully!")
    print(f"📅 Date: {today}")
    print(f"💾 Saved: {EXCEL_FILE}\n")

except FileNotFoundError:
    print(f"\n❌ Could not find {EXCEL_FILE}")
    print("Make sure it's in the same folder as this script.\n")

except Exception as e:
    print(f"\n❌ Something went wrong: {str(e)}\n")
